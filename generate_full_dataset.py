import csv
import argparse
import os
import sys
import time
from typing import Dict, List, Tuple
from playwright.sync_api import Page, sync_playwright
from playwright_stealth import stealth_sync
from openpyxl import Workbook, load_workbook
import json
from datetime import datetime
import traceback


# Configuration
CONFIG = {
    "input_files": {
        "cars": "initial_dataset/cars_makes_and_years.csv",
        "rvs": "initial_dataset/rvs_makes_and_years.csv",
        "boats": "initial_dataset/boats_makes_and_years.csv",
        "motorcycles": "initial_dataset/motorcycles_makes_and_years.csv",
    },
    "output_file": "full_dataset/vehicle_data.xlsx",
    "base_urls": {
        "cars": "https://www.jdpower.com/cars/{year}/{make}",
        "rvs": "https://www.jdpower.com/rvs/{year}/{make}",
        "boats": "https://www.jdpower.com/boats/{year}/{make}",
        "motorcycles": "https://www.jdpower.com/motorcycles/{year}/{make}",
    },
    "headers": {
        "cars": ["Year", "Vehicle Type", "Make", "Model", "Trim", "Blurb"],
        "rvs": ["Year", "Vehicle Type", "Make", "Model", "Trim", "Blurb"],
        "boats": ["Year", "Vehicle Type", "Make", "Model", "Length", "Model Type", 
                 "Hull", "CC's", "Engine(s)", "HP", "Weight (lbs)", "Fuel Type", "Blurb"],
        "motorcycles": ["Year", "Vehicle Type", "Make", "Model", "Trim", "Blurb"],
    }
}


def sanitize_make(make: str) -> str:
    """Replace spaces and slashes with hyphens and convert to lowercase."""
    return make.replace(' ', '-').replace('/', '-').lower()

class CheckpointManager:
    def __init__(self, checkpoint_file="checkpoint.json"):
        self.checkpoint_file = checkpoint_file
        self.state = {
            'current_vehicle_type': None,
            'current_make': None,
            'processed_years': {},
            'error_log': []
        }
        self.load()

    def load(self):
        if os.path.exists(self.checkpoint_file):
            try:
                with open(self.checkpoint_file, 'r') as f:
                    self.state = json.load(f)
            except Exception as e:
                print(f"Error loading checkpoint: {e}. Starting fresh.")

    def save(self):
        with open(self.checkpoint_file, 'w') as f:
            json.dump(self.state, f)

    def log_error(self, error_info):
        self.state['error_log'].append({
            'timestamp': datetime.now().isoformat(),
            'error': error_info
        })
        self.save()

    def update_progress(self, vehicle_type, make, year):
        key = f"{vehicle_type}-{make}"
        if key not in self.state['processed_years']:
            self.state['processed_years'][key] = []
        if year not in self.state['processed_years'][key]:
            self.state['processed_years'][key].append(year)
        self.save()

    def should_process(self, vehicle_type, make, year):
        key = f"{vehicle_type}-{make}"
        return year not in self.state['processed_years'].get(key, [])

class ErrorHandler:
    @staticmethod
    def handle_error(checkpoint, error, context=None):
        error_info = {
            'timestamp': datetime.now().isoformat(),
            'error': str(error),
            'context': context,
            'traceback': traceback.format_exc()
        }
        checkpoint.log_error(error_info)
        print(f"Error occurred: {error}")
        print(f"Context: {context}")
        print("Checkpoint saved. Restart script to resume.")


class ExcelManager:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.workbook = self._initialize_workbook()
        self.sheets = {}

    def _initialize_workbook(self) -> Workbook:
        if os.path.exists(self.output_path):
            try:
                return load_workbook(self.output_path)
            except Exception as e:
                print(f"Error loading workbook: {e}. Creating new workbook.")
        return Workbook()

    def get_sheet(self, vehicle_type: str):
        if vehicle_type not in self.sheets:
            sheet_name = vehicle_type.capitalize()
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
            else:
                sheet = self.workbook.create_sheet(title=sheet_name)
                sheet.append(CONFIG["headers"][vehicle_type])
            self.sheets[vehicle_type] = sheet
        return self.sheets[vehicle_type]

    def save(self):
        self.workbook.save(self.output_path)

class BrowserManager:
    def __init__(self):
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.firefox.launch(headless=True)
        self.context = self.browser.new_context(ignore_https_errors=True)
        self.page = self.context.new_page()
        stealth_sync(self.page)

    def __enter__(self):
        return self.page

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.context.close()
        self.browser.close()
        self.playwright.stop()

class BaseScraper:
    def __init__(self, excel_manager: ExcelManager, vehicle_type: str):
        self.excel = excel_manager
        self.vehicle_type = vehicle_type
        self.sheet = self.excel.get_sheet(vehicle_type)

    def process_make(self, make: str, years: List[str], selected_years: List[str]):
        raise NotImplementedError

    @staticmethod
    def read_csv(file_path: str) -> List[Tuple[str, List[str]]]:
        makes = []
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)  # Skip header
            for row in reader:
                makes.append((row[0], row[1].split(", ")))
        return makes

class CarScraper(BaseScraper):
    def process_make(self, make: str, years: List[str], selected_years: List[str]):
        for year in selected_years:
            if year not in years:
                continue
            with BrowserManager() as page:
                self._process_year(make, year, page)

    def _process_year(self, make: str, year: str, page: Page):
        sanitized_make = sanitize_make(make)  # Sanitize the make
        url = CONFIG["base_urls"]["cars"].format(year=year, make = sanitized_make)
        print(url)

        page.goto(url, timeout=60000)
        time.sleep(5)
        
        model_elements = page.query_selector_all(".yearMake_model-wrapper-h3__npC2B h3")
        for model_element in model_elements:
            model_name = model_element.inner_text().strip()
            print(f"Fetching trims for model: {model_name}...")
            self._process_model(page, model_element, year, make, model_name)


    def _process_model(self, page: Page, model_element, year: str, make: str, model_name: str):
        model_url = model_element.evaluate("node => node.closest('.yearMake_model-wrapper__t8GAv').querySelector('a').href")
        
        with page.context.expect_page() as new_tab_info:
            page.evaluate(f"window.open('{model_url}', '_blank')")
        new_tab = new_tab_info.value
        
        try:
            new_tab.wait_for_selector(".trimSelection_card-info__O02As", timeout=60000)
            trim_containers = new_tab.query_selector_all(
                ".MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-md-6.trimSelection_card-info__O02As"
            )
            
            for trim_container in trim_containers:
                # Locate the trim name header
                trim_name_element = trim_container.query_selector("h3.heading-xs.title.spacing-s")
                #model_name = trim_name_element.inner_text().strip() if trim_name_element else "Unknown Model"

                # Locate all trims under the model
                trim_links = trim_container.query_selector_all(
                    ".MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-md-12 a"
                )
                for trim_link in trim_links:
                    trim_name = trim_link.inner_text().strip()
                    print(year, "cars", make, model_name, trim_name)

                    self.sheet.append([year, "cars", make, model_name, trim_name])
                    self.excel.save()
                
        finally:
            new_tab.close()

class RVScraper(BaseScraper):
    def process_make(self, make: str, years: List[str], selected_years: List[str]):
        for year in selected_years:
            if year not in years:
                continue
            with BrowserManager() as page:
                self._process_year(make, year, page)

    def _process_year(self, make: str, year: str, page: Page):
        sanitized_make = sanitize_make(make)  # Sanitize the make
        url = CONFIG["base_urls"]["rvs"].format(year=year, make=sanitized_make)
        print(url)

        page.goto(url, timeout=60000)
        
        # Wait for the RV table to load
        page.wait_for_selector("table.table-enhanced--model-years", timeout=60000)

        rows = page.query_selector_all("table.table-enhanced--model-years tr")
        current_model = None
        consecutive_tr_count = 0

        for row in rows:
            # Handle consecutive <tr> tags for models
            if not row.get_attribute("class"):
                consecutive_tr_count += 1
                if consecutive_tr_count == 2:  # Second consecutive <tr> is the model
                    model_header = row.query_selector("td[colspan='6'] h4")
                    if model_header:
                        current_model = model_header.inner_text().strip()
                        print(f"Found model (from consecutive <tr>): {current_model}")
                    consecutive_tr_count = 0
                continue
            else:
                consecutive_tr_count = 0

            # Check for <th> headers as models
            th_header = row.query_selector("th h3.category")
            if th_header:
                current_model = th_header.inner_text().strip()
                print(f"Found model (from <th>): {current_model}")
                continue

            # Check if the row defines a trim
            if row.get_attribute("class") == "detail-row":
                trim_element = row.query_selector("td a")
                if trim_element and current_model:
                    trim_name = trim_element.inner_text().strip()
                    print(f"Found trim: {trim_name} for model: {current_model}")
                    # Append to the XLSX sheet
                    print(year, "rvs", make, current_model, trim_name)

                    self.sheet.append([year, "rvs", make, current_model, trim_name])
                    self.excel.save()

class BoatScraper(BaseScraper):
    def process_make(self, make: str, years: List[str], selected_years: List[str]):
        for year in selected_years:
            if year not in years:
                continue
            with BrowserManager() as page:
                self._process_year(make, year, page)

    def _process_year(self, make: str, year: str, page: Page):
        sanitized_make = sanitize_make(make)  # Sanitize the make
        url = CONFIG["base_urls"]["boats"].format(year=year, make=sanitized_make)
        print(f"Processing URL: {url}")

        page.goto(url, timeout=60000)
        
        # Wait for the main content container
        if page.wait_for_selector(".MuiGrid-container", timeout=15000):
            # Extract all rows with complete data
            rows = page.query_selector_all(".MuiGrid-root.MuiGrid-item.MuiGrid-grid-md-12.mui-190ub4r")
            
            for row in rows:
                # Check if the row contains all the required data
                columns = row.query_selector_all(".MuiGrid-root.MuiGrid-item")
                if len(columns) == 9:  # Ensure there are 9 columns (Model, Length, Model Type, Hull, CC's, Engine(s), HP, Weight (lbs), Fuel Type)
                    # Extract the data from each column
                    model = columns[0].inner_text().strip()
                    length = columns[1].inner_text().strip()
                    model_type = columns[2].inner_text().strip()
                    hull = columns[3].inner_text().strip()
                    ccs = columns[4].inner_text().strip()
                    engines = columns[5].inner_text().strip()
                    hp = columns[6].inner_text().strip()
                    weight = columns[7].inner_text().strip()
                    fuel_type = columns[8].inner_text().strip()

                    # Append the data to the Excel sheet
                    self.sheet.append([
                        year, "boat", make, model, length, model_type, hull, ccs, engines, hp, weight, fuel_type
                    ])
                    self.excel.save()
                    print(f"Appended row: {[year, 'boat', make, model, length, model_type, hull, ccs, engines, hp, weight, fuel_type]}")


class MotorcycleScraper(BaseScraper):
    def process_make(self, make: str, years: List[str], selected_years: List[str]):
        for year in selected_years:
            if year not in years:
                continue
            with BrowserManager() as page:
                self._process_year(make, year, page)

    def _process_year(self, make: str, year: str, page: Page):
        sanitized_make = sanitize_make(make)  # Sanitize the make
        url = CONFIG["base_urls"]["motorcycles"].format(year=year, make=sanitized_make)
        page.goto(url, timeout=60000)
        
        page.wait_for_selector(".spacing-xs h3.heading-s", timeout=60000)
        sections = page.query_selector_all(".spacing-xs + .spacing-s")  # Select the second `.spacing-s` div

        for section in sections:
            model_element = section.query_selector("h4.bh-l")
            if not model_element:
                continue

            model_name = model_element.inner_text().strip()
            print(f"Processing model: {model_name}")

            # Fetch trims under the current model
            trims = section.query_selector_all(
                ".motorcyclesYearMake_model-link-container__JIYG4 a.motorcyclesYearMake_model-link__Db22K"
            )

            for trim_element in trims:
                trim_name = trim_element.inner_text().strip()
                print(f"Found trim: {trim_name} for model: {model_name}")
                self.sheet.append([year, "motorcycles", make, model_name, trim_name])
                self.excel.save()


def parse_arguments():
    parser = argparse.ArgumentParser(description="Scrape vehicle data from JDPower.")
    parser.add_argument("--years", type=str, required=True, help="Year or year range")
    parser.add_argument("-c", action="store_true", help="Process cars")
    parser.add_argument("-r", action="store_true", help="Process RVs")
    parser.add_argument("-b", action="store_true", help="Process boats")
    parser.add_argument("-m", action="store_true", help="Process motorcycles")
    parser.add_argument("-all", action="store_true", help="Process all vehicle types")
    return parser.parse_args()

def process_arguments(args) -> Tuple[List[str], List[str]]:
    if "-" in args.years:
        start, end = map(int, args.years.split("-"))
        years = list(map(str, range(start, end + 1)))
    else:
        years = [args.years]
    
    types = []
    if args.all:
        types = ["cars", "rvs", "boats", "motorcycles"]
    else:
        if args.c: types.append("cars")
        if args.r: types.append("rvs")
        if args.b: types.append("boats")
        if args.m: types.append("motorcycles")
    
    if not types:
        print("No vehicle types selected!")
        sys.exit(1)
    
    return years, types

def main():
    args = parse_arguments()
    selected_years, selected_types = process_arguments(args)
    
    checkpoint = CheckpointManager()
    excel_manager = ExcelManager(CONFIG["output_file"])
    
    scraper_map = {
        "cars": CarScraper(excel_manager, "cars"),
        "rvs": RVScraper(excel_manager, "rvs"),
        "boats": BoatScraper(excel_manager, "boats"),
        "motorcycles": MotorcycleScraper(excel_manager, "motorcycles")
    }
    
    try:
        for vehicle_type in selected_types:
            scraper = scraper_map[vehicle_type]
            makes = scraper.read_csv(CONFIG["input_files"][vehicle_type])
            
            for make, years in makes:
                try:
                    for year in selected_years:
                        if year not in years or not checkpoint.should_process(vehicle_type, make, year):
                            continue
                        
                        retries = 10
                        while retries > 0:
                            try:
                                scraper.process_make(make, years, [year])
                                checkpoint.update_progress(vehicle_type, make, year)
                                break
                            except Exception as e:
                                retries -= 1
                                if retries == 0:
                                    raise
                                print(f"Retrying {vehicle_type}/{make}/{year} ({retries} left)...")
                                time.sleep(60)  # Wait before retrying
                                
                except Exception as e:
                    ErrorHandler.handle_error(
                        checkpoint, e,
                        context=f"{vehicle_type}/{make}"
                    )
                    continue  # Continue with next make
    
    except KeyboardInterrupt:
        print("\nKeyboard interrupt received. Saving checkpoint...")
        checkpoint.save()
        sys.exit(0)
        
    except Exception as e:
        ErrorHandler.handle_error(checkpoint, e)
        sys.exit(1)

if __name__ == "__main__":
    main()
